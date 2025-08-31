from openpyxl import load_workbook
import webbrowser
import os

# Nomes dos arquivos Excel
excel_files = {
    "mundo1": "mundo1.xlsx",
    "mundo2": "mundo2.xlsx",
    "mundo3": "mundo3.xlsx"
}

# Nome do arquivo HTML que será gerado
html_file = "index.html"

# Começo do HTML
html_start = """<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Curso em Video Exercises</title>
<style>
  body { font-family: Arial, sans-serif; padding: 20px; }
  h1 { text-align: center; margin-bottom: 10px; }
  .tabs { display: flex; cursor: pointer; margin-bottom: 20px; }
  .tab { padding: 10px 20px; border: 1px solid #ccc; border-bottom: none; }
  .tab.active { background-color: #f0f0f0; font-weight: bold; }
  .tab-content { border: 1px solid #ccc; padding: 20px; display: none; }
  .tab-content.active { display: block; }
  pre { background-color: #eee; padding: 10px; border-radius: 5px; position: relative; white-space: pre-wrap; word-wrap: break-word; }
  .copy-btn { position: absolute; top: 5px; right: 5px; padding: 5px 10px; font-size: 12px; cursor: pointer; }
  .exercise { position: relative; margin-bottom: 20px; }
  p.description { text-align: center; font-size: 16px; color: #555; margin-bottom: 30px; }
</style>
</head>
<body>

<h1>Curso em Video Exercises</h1>

<p class="description">
<b>Português:</b> Este site organiza os exercícios do Curso em Vídeo em três mundos. No Mundo 1, você encontra os vídeos do curso com título, link e a primeira frase da descrição. Cada exercício possui um botão de copiar para facilitar o estudo e a referência rápida. Mundos 2 e 3 estão prontos para receber os próximos exercícios à medida que forem criados.
</p>

<p class="description">
<b>English:</b> This website organizes the exercises from the Curso em Vídeo into three worlds. In World 1, you can find the course videos with their title, link, and the first sentence of the description. Each exercise has its own copy button to make studying and referencing easier. Worlds 2 and 3 are ready to receive future exercises as they are created.
</p>

<div class="tabs">
  <div class="tab active" data-tab="mundo1">Mundo 1</div>
  <div class="tab" data-tab="mundo2">Mundo 2</div>
  <div class="tab" data-tab="mundo3">Mundo 3</div>
</div>
"""

html_end_template = """
<div id="{id}" class="tab-content{active}">
{content}
</div>
"""

html_after_tabs = """
<script>
  // Tabs
  const tabs = document.querySelectorAll('.tab');
  const contents = document.querySelectorAll('.tab-content');
  tabs.forEach(tab => {
    tab.addEventListener('click', () => {
      tabs.forEach(t => t.classList.remove('active'));
      tab.classList.add('active');
      contents.forEach(c => c.classList.remove('active'));
      document.getElementById(tab.dataset.tab).classList.add('active');
    });
  });

  // Copy buttons
  document.querySelectorAll('.exercise').forEach(exercise => {
    const btn = exercise.querySelector('.copy-btn');
    const pre = exercise.querySelector('pre');
    btn.addEventListener('click', () => {
      navigator.clipboard.writeText(pre.innerText.trim()).then(() => {
        btn.textContent = 'Copiado!';
        setTimeout(() => btn.textContent = 'Copiar', 1500);
      });
    });
  });
</script>
</body>
</html>
"""

def generate_exercises_html(file_path):
    """Return HTML blocks for exercises from a given Excel file."""
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
    except FileNotFoundError:
        return None  # Arquivo não existe

    exercises_html = ""
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        title, link, description = row[:3]
        first_sentence = description.split('.')[0].strip() if description else ""
        block = f"""
  <div class="exercise">
    <pre>
# Título do vídeo: {title}
#
# Link do vídeo: {link}
# Descrição: {first_sentence}
    </pre>
    <button class="copy-btn">Copiar</button>
  </div>
"""
        exercises_html += block
    return exercises_html if exercises_html else None

# Montar conteúdo de cada mundo
worlds_content = {}
for world_id, file_name in excel_files.items():
    html_blocks = generate_exercises_html(file_name)
    if html_blocks:
        content = html_blocks
    else:
        content = f"<p>Conteúdo do {world_id.replace('mundo','Mundo ')} aqui...</p>"
    active_class = " active" if world_id == "mundo1" else ""
    worlds_content[world_id] = html_end_template.format(id=world_id, content=content, active=active_class)

# Montar HTML final
html_final = html_start + "".join(worlds_content.values()) + html_after_tabs

# Salvar HTML
with open(html_file, "w", encoding="utf-8") as f:
    f.write(html_final)

print(f"HTML gerado com sucesso: {html_file}")

# Abrir automaticamente no navegador
html_path = os.path.abspath(html_file)
webbrowser.open(f"file://{html_path}")
